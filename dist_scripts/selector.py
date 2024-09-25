from enum import Enum, unique
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles.numbers import FORMAT_TEXT
from acad_extractor import select_on_screen
import re
import win32com.client
from pythoncom import VT_ARRAY, VT_R8
import json
from pyproj import Proj


def to_vr8(x):
    return win32com.client.VARIANT(VT_ARRAY | VT_R8, x)


def zoom(h):
    app = win32com.client.Dispatch("AutoCAD.Application")
    doc = app.ActiveDocument
    obj = doc.HandleToObject(h)
    bb = obj.GetBoundingBox()
    app.ZoomWindow(to_vr8(bb[0]), to_vr8(bb[1]))


R1C1_to_A1_RE = re.compile("R\[(-*\d+)\]C\[(-*\d+)\]")


def R1C1_to_A1(row, col, s):
    return re.sub(
        R1C1_to_A1_RE,
        lambda match: f"{get_column_letter(col + int(match.group(2)))}{row + int(match.group(1))}",
        s
    )


@unique
class Cols(Enum):
    # col, precision
    HANDLE = 1, None
    VIS = 2, None
    NAME = 3, None
    CODE = 4, None
    NUMBER = 5, None
    ADDR = 6, None
    LEVELS = 7, None
    BLOCKS = 8, None
    APARTS = 9, None
    MARKA = 10, None
    L = 11, 1
    LE = 12, 1
    MTK = 13, None
    MTK_L = 14, 1
    MTK_LS = 15, 1
    MTK_F = 16, None
    MTK_T = 17, None
    MTK_LE = 18, 1
    DC_L = 19, 1
    DC_LE = 20, 1
    NOMRE_1 = 21, None
    NOMRE_2 = 22, None
    KOL = 23, None
    COLOR = 24, None
    X = 25, 4
    Y = 26, 4
    LAT = 27, 6
    LON = 28, 6


def create_name_maps(x):
    acad_to_app = {}
    for app_name, acad_names in x.items():
        for acad_name in acad_names:
            acad_to_app[acad_name] = app_name
    return acad_to_app


ATTRS = create_name_maps({
    # app_name -> acad_names
    "CODE": ("КОД", "KOD"),
    "NUMBER": ("N_ДОМА", "N_BİNA"),
    "ADDR": ("АДРЕС", "ÜNVAN"),
    "LEVELS": ("ЭТАЖ", "MƏRTƏBƏLƏRIN_SAYI"),
    "APARTS": ("ВСЕГО_КВАР", "MƏNZILLƏRIN_ÜMUMI_SAYI"),
    "MARKA": ("FO_MARKA",),
    "LE": ("ƏLAVƏ_YERALTİ_KABEL",),
    "MTK": ("FO_MTK",),
    "MTK_LS": ("L_ДО_ШАХТЫ", "ŞAXTAYA_QƏDƏR_MƏSAFƏ"),
    "MTK_F": ("КОЛ_ЭТ_ОТ", "MƏRTƏBƏDƏN"),
    "MTK_T": ("КОЛ_ЭТ_ДО", "MƏRTƏBƏYƏ_KİMİ"),
    "MTK_LE": ("L_ЗАПАС", "KABEL_EHTİYATİ"),
    "DC_LE": ("L-ДОП",),
    "NOMRE_1": ("NOMRE_1",),
    "NOMRE_2": ("NOMRE_2",),
    "KOL": ("KOL",),
    "COLOR": ("COLOR",),
})

PROPS = create_name_maps({
    # app_name -> acad_names
    "VIS": ("Visibility1",),
})

NAMES = create_name_maps({
    # app_name -> acad_names
    ".Dom": (".Dom", ".BINA"),
})

EXCLUDE = set([
    "Port",
    "Strelka_V",
    "Vereg"
])


def select(callback=None, utm_zone=None):
    items = []

    names = []
    attrs_list = []
    props_list = []
    handles = []
    points = []
    select_on_screen(names, attrs_list, props_list, handles, callback, points)

    proj = None if utm_zone is None else Proj(proj="utm", zone=utm_zone, ellps="WGS84", preserve_units=False)

    for i, name in enumerate(names):
        name_ = NAMES.get(name) or name
        if name_ in EXCLUDE:
            continue

        item = {}

        item["NAME"] = name
        item["HANDLE"] = handles[i]
        item["X"] = points[i][0]
        item["Y"] = points[i][1]
        if utm_zone is not None:
            lon, lat = proj(item["X"], item["Y"], inverse=True)
            item["LAT"] = lat
            item["LON"] = lon

        attrs = {}
        for attr_name, value in attrs_list[i]:
            attrs[attr_name] = value

        props = {}
        for prop_name, value in props_list[i]:
            props[prop_name] = value

        data = {}
        for src, MAP in [(attrs, ATTRS), (props, PROPS)]:
            for acad_name, value in src.items():
                app_name = MAP.get(acad_name)
                if app_name is not None:
                    data[app_name] = value

        item["CODE"] = data.get("CODE", "")
        item["VIS"] = data.get("VIS")

        if name_ == ".Dom":
            item["NUMBER"] = data["NUMBER"]
            item["ADDR"] = data["ADDR"]
            item["LEVELS"] = int(data["LEVELS"])
            item["BLOCKS"] = int(data["VIS"])
            item["APARTS"] = int(data["APARTS"])
        elif name_ == ".FO_cabel":
            s = 0
            for index in range(int(data["VIS"])):
                s += props[f"L_{index + 1:02d}"]
            item["MARKA"] = data["MARKA"]
            item["LE"] = float(data.get("LE", "0"))
            L_f = f"={s}+20+R[0]C[1]"
            L_v = s + 20 + item["LE"]
            item["L"] = L_v, L_f
        elif name_ == ".FO_MTK":
            item["MTK"] = data["MTK"]
            item["MTK_LS"] = float(data["MTK_LS"])
            item["MTK_F"] = int(data["MTK_F"])
            item["MTK_T"] = int(data["MTK_T"])
            item["MTK_LE"] = float(data["MTK_LE"])
            MTK_L_f = "=R[0]C[1]+ABS(R[0]C[2]-R[0]C[3])*4+R[0]C[4]"
            MTK_L_v = item["MTK_LS"] + abs(item["MTK_F"] - item["MTK_T"]) * 4 + item["MTK_LE"]
            item["MTK_L"] = MTK_L_v, MTK_L_f
        elif name_ == ".FO2_DC":
            item["DC_L"] = int(data["VIS"])
            item["DC_LE"] = int(data["DC_LE"])
        elif name_ == "KOL_DD":
            item["KOL"] = int(data["KOL"])
            item["COLOR"] = data["COLOR"]
        for k in ["NOMRE_1", "NOMRE_2"]:
            v = data.get(k)
            if k is not None:
                item[k] = v

        items.append(item)

    items.sort(key=lambda item: (item["CODE"], item["NAME"]))
    return items


# MTEXT Format Codes
# https://www.cadforum.cz/en/text-formatting-codes-in-mtext-objects-tip8640
# https://adndevblog.typepad.com/autocad/2017/09/dissecting-mtext-format-codes.html
FORMAT_RE = re.compile("(\\\\[LlOoKkNPX~]|\\\\[QHWfACT].*?;|\{|\})")

v12345 = "\\A1;{\\fCourier New|b0|i0|c204|p49;\\~\\L1\\l2\\O3\\o4\\fCourier New|b1|i0|c204|p49;5}"
v67 = "\\A1;{\\fCourier New|b0|i1|c204|p49;\\P6\\fCourier New|b0|i0|c204|p49;\\K7}"


def unformat(s):
    return re.sub(FORMAT_RE, "", s)


def select_and_sum(callback=None):
    try:
        app = win32com.client.Dispatch("AutoCAD.Application")
    except:
        raise RuntimeError("no app")

    if not app.Visible:
        app.Quit()
        raise RuntimeError("app.Visible == False")

    try:
        doc = app.ActiveDocument
    except:
        raise RuntimeError("no doc")

    if doc is None:
        raise RuntimeError("doc is None")

    try:
        selection = doc.ActiveSelectionSet
        selection.Clear()
        selection.SelectOnScreen()
    except:
        raise RuntimeError("selection error")

    s = []
    count = len(selection)
    callback_delta = int(count / 100)
    if callback_delta == 0:
        callback_delta = 1
    for i, item in enumerate(selection):
        if callback is not None and i % callback_delta == 0:
            callback(i, count)
        if item.EntityName == "AcDbMText":
            s.append(int(unformat(item.textString.strip())))
    if callback is not None:
        callback(count, count)
    return s


def write_blocks(wb, items):
    ws = wb["BLOCKS"]
    for col_desc in Cols:
        col, _ = col_desc.value
        ws.cell(1, col, col_desc.name)

    for row, item in enumerate(items):
        for col_desc in Cols:
            col, precision = col_desc.value
            value = item.get(col_desc.name)
            if value is not None:
                if isinstance(value, tuple):
                    value = R1C1_to_A1(row + 2, col, value[1])
                cell = ws.cell(row + 2, col, value)
                if precision is not None:
                    cell.number_format = f"0.{'0' * precision}"
                elif not isinstance(value, (float, int)):
                    cell.number_format = FORMAT_TEXT


if __name__ == "__main__":
    assert unformat(v12345) == "12345"
    assert unformat(v67) == "67"

    def callback(processed, total):
        print(f"Processed {processed}/{total}")
    items = select(callback, 39)

    wb = openpyxl.Workbook()
    wb.active.title = "BLOCKS"
    write_blocks(wb, items)
    wb.save("test.xlsx")
    wb.close()

    with open("test.json", "w") as f:
        f.write(json.dumps(items))
