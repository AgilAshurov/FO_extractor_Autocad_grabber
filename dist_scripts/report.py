import openpyxl
import re
import json
import io


def m_to_re(m):
    s = "^"
    n = 0
    for c in m:
        if c == "A":
            n += 1
        else:
            if n:
                s += f"[A-Za-z]{{{n}}}"
                n = 0
            if c == "N":
                if s.endswith("\\d+"):
                    raise RuntimeError(f"bad mask {m}")
                s += "\\d+"
            elif c == "*":
                s += ".*"
            else:
                s += c
    if n:
        s += f"[A-Za-z]{{{n}}}"
    s += "$"
    return re.compile(s)


class Query:
    def __init__(self, props, row, col):
        self.row = row
        self.col = col

        self.f = props.pop("f", None)
        if self.f is not None:
            self.f = self.f[0].split("+")

        self.d = props.pop("d", None)
        if self.d is not None:
            self.d = int(self.d[0])

        self.v_props = {}
        self.m_props = {}
        for prop, value in props.items():
            mask = prop.endswith("m")
            if not mask:
                self.v_props[prop] = value
            else:
                self.m_props[prop[:-1]] = [m_to_re(e) for e in value]

        self.value = 0

    @staticmethod
    def add(qs, row, col, s):
        if not isinstance(s, str) or not s.startswith(":"):
            return
        props = {}
        items = s[1:].split(":")
        for item in items:
            subitems = item.split("=")
            if len(subitems) < 2:
                raise RuntimeError(f"bad query {s}")
            props[subitems[0]] = subitems[1:]
        qs.append(Query(props, row, col))

    def sum(self, item):
        for prop, value in self.v_props.items():
            x = item.get(prop)
            if x is None:
                return False
            if not any(x == e for e in value):
                return False

        for prop, value in self.m_props.items():
            x = item.get(prop)
            if x is None:
                return False
            if not any(e.match(x) is not None for e in value):
                return False

        if self.f is not None:
            for key in self.f:
                self.value += item[key][0] if isinstance(item[key], (list, tuple)) else item[key]
        else:
            self.value += 1

        return True


def write_report(wb, items):
    ws = wb["REPORT"]
    qs = []
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            Query.add(qs, row, col, ws.cell(row, col).value)
    for item in items:
        for q in qs:
            q.sum(item)
    for q in qs:
        cell = ws.cell(q.row, q.col, q.value)
        if q.d is not None:
            cell.number_format = f"0.{'0' * q.d}"


if __name__ == "__main__":
    m = m_to_re("AN")
    assert m.pattern == "^[A-Za-z]{1}\\d+$"

    m = m_to_re("AAN-N")
    assert m.pattern == "^[A-Za-z]{2}\\d+-\\d+$"

    m = m_to_re("FOAN")
    assert m.pattern == "^FO[A-Za-z]{1}\\d+$"

    m = m_to_re("FO*")
    assert m.pattern == "^FO.*$"

    qs = []
    Query.add(qs, 1, 2, ":NAME=NAME1:COLOR=red=blue:CODEm=AN:f=X+Y:d=2")
    q = qs[0]
    assert q.row == 1
    assert q.col == 2
    assert q.f == ["X", "Y"]
    assert q.d == 2
    assert len(q.v_props) == 2
    assert len(q.m_props) == 1
    assert q.v_props["NAME"] == ["NAME1"]
    assert q.v_props["COLOR"] == ["red", "blue"]
    assert len(q.m_props["CODE"]) == 1
    assert q.m_props["CODE"][0].pattern == "^[A-Za-z]{1}\\d+$"

    q.sum({"NAME": "NAME1", "COLOR": "red", "CODE": "X1", "X": 1, "Y": 2})
    q.sum({"NAME": "NAME1", "COLOR": "blue", "CODE": "Y234", "X": 10, "Y": 20})
    q.sum({"NAME": "NAME2", "COLOR": "red", "CODE": "X1", "X": 100, "Y": 200})
    q.sum({"NAME": "NAME1", "COLOR": "green", "CODE": "X1", "X": 100, "Y": 200})
    q.sum({"NAME": "NAME1", "COLOR": "red", "CODE": "234", "X": 100, "Y": 200})
    assert q.value == 33

    qs = []
    Query.add(qs, 1, 2, ":MARKAm=FO8AN-N:f=L")
    q = qs[0]
    q.sum({"MARKA": "FO8A9-2", "L": 100})
    assert q.value == 100

    qs = []
    Query.add(qs, 1, 2, ":MARKAm=FO8*:f=L")
    q = qs[0]
    q.sum({"MARKA": "FO8A9", "L": 100})
    assert q.value == 100

    if False:
        with open("test.json", "r") as f:
            items = json.loads(f.read())
        with open("test_report.xlsx", "rb") as f:
            content = f.read()
        wb = openpyxl.load_workbook(io.BytesIO(content))
        write_report(wb, items)
        wb.save("test.xlsx")
        wb.close()
